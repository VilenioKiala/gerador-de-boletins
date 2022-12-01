import openpyxl,os,datetime;

def criarBoletim(dados):
	created_at = str(datetime.datetime.now()).split(" ")
	created_at = "-".join(created_at)
	created_at = created_at.split(":")
	created_at = "-".join(created_at)
	created_at = created_at.split(".")
	created_at = "-".join(created_at)


	qnt_nota = len(dados['caderneta'])

	wb = openpyxl.Workbook()
	ws = wb.active
	ws["A1"] = "República de Angola"
	ws["A2"] = "Nome do Aluno: " + dados['aluno'].title()
	ws["A3"] = ("Classe: " + str(dados["classe"]) + "      " + "Trimestre: " + 
		str(dados["trimestre"])+"º"+ "         " + "Ano: " + str(dados["ano_lectivo"]))
	ws["A4"] = ("Media: " + str(dados["media"]) + "        " + "Situação: " + 
		dados["situacao"].title()
	)

	ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=(int(qnt_nota)-1))
	ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=(int(qnt_nota)-1))
	ws.merge_cells(start_row=3,end_row=3,start_column=1,end_column=(int(qnt_nota)-1))
	ws.merge_cells(start_row=4,end_row=4,start_column=1,end_column=(int(qnt_nota)-1))

	for i in range(0,(qnt_nota)):
		ws.cell(row=5,column=(i+1),value=dados['caderneta'][i-1]['disciplina'].upper())

	for j in range(0,(qnt_nota)):
		ws.cell(row=6,column=(j+1),value=dados['caderneta'][j-1]["nota"])


	filename = "boletim" + created_at + ".xlsx"
	wb.save(filename);
	
	print("Boletim gerado com sucesso!")
